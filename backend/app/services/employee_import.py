"""استيراد بيانات الموظفين من ملف Excel، وإنشاء نموذج فارغ للتنزيل.

يدعم عناوين أعمدة عربية أو إنجليزية شائعة، بلا حاجة لصيغة صارمة واحدة.
"""
from __future__ import annotations

from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook, load_workbook

COLUMN_ALIASES: dict[str, list[str]] = {
    "full_name": ["الاسم", "اسم الموظف", "الاسم الكامل", "name", "full name", "full_name"],
    "employee_number": ["الرقم الوظيفي", "رقم الموظف", "employee number", "employee_number", "employee id"],
    "nationality": ["الجنسية", "nationality"],
    "job_title": ["المسمى الوظيفي", "الوظيفة", "job title", "job_title", "position"],
    "phone": ["رقم الجوال", "الجوال", "phone", "mobile"],
    "salary": ["الراتب", "salary"],
    "iqama_number": ["رقم الإقامة", "رقم الاقامة", "iqama number", "iqama_number", "iqama"],
    "iqama_expiry_date": [
        "تاريخ انتهاء الإقامة",
        "تاريخ انتهاء الاقامة",
        "تاريخ الانتهاء",
        "iqama expiry",
        "iqama_expiry_date",
        "expiry date",
        "expiry_date",
    ],
    "national_id": ["رقم الهوية", "رقم الهوية الوطنية", "national id", "national_id"],
    "hire_date": ["تاريخ التعيين", "تاريخ الالتحاق", "hire date", "hire_date", "start date"],
    "basic_salary": ["الراتب الأساسي", "الراتب الاساسي", "basic salary", "basic_salary"],
    "housing_allowance": ["بدل السكن", "بدل سكن", "housing allowance", "housing_allowance"],
    "other_allowances": ["بدلات أخرى", "بدلات اخرى", "other allowances", "other_allowances"],
    "contract_type": ["نوع العقد", "contract type", "contract_type"],
    "contract_end_date": ["تاريخ نهاية العقد", "contract end date", "contract_end_date"],
    "probation_end_date": ["تاريخ نهاية التجربة", "probation end date", "probation_end_date"],
}

TEMPLATE_HEADERS_AR = [
    "الاسم", "الرقم الوظيفي", "الجنسية", "المسمى الوظيفي",
    "رقم الجوال", "الراتب", "رقم الإقامة", "تاريخ انتهاء الإقامة",
    "رقم الهوية", "تاريخ التعيين", "الراتب الأساسي", "بدل السكن",
    "بدلات أخرى", "نوع العقد", "تاريخ نهاية العقد", "تاريخ نهاية التجربة",
]


class EmployeeImportError(Exception):
    pass


def _normalize_header(value) -> str:
    return str(value).strip().lower() if value is not None else ""


def _clean_str(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _match_columns(header_row: tuple) -> dict[str, int]:
    normalized = [_normalize_header(h) for h in header_row]
    mapping: dict[str, int] = {}
    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias.strip().lower() in normalized:
                mapping[field] = normalized.index(alias.strip().lower())
                break
    return mapping


def _parse_date_cell(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip() or None


# اسم قديم يُستخدم أيضًا لتواريخ الإقامة تحديدًا (تم إبقاؤه للتوافق العكسي)
_parse_expiry = _parse_date_cell


def _parse_number(value) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_employees_excel(content: bytes) -> tuple[list[dict], list[dict]]:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception as e:
        raise EmployeeImportError(f"تعذّر قراءة ملف Excel: {e}")

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise EmployeeImportError("الملف فارغ.")

    mapping = _match_columns(rows[0])
    if "full_name" not in mapping:
        raise EmployeeImportError(
            "لم يتم العثور على عمود الاسم. تأكد من وجود عمود بعنوان 'الاسم' أو 'name'."
        )

    def get(row: tuple, field: str):
        idx = mapping.get(field)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    employees: list[dict] = []
    errors: list[dict] = []

    for row_num, row in enumerate(rows[1:], start=2):
        if row is None or all(cell is None for cell in row):
            continue

        full_name = get(row, "full_name")
        if not full_name or not str(full_name).strip():
            errors.append({"row": row_num, "reason": "لا يوجد اسم في هذا الصف"})
            continue

        contract_type_raw = _clean_str(get(row, "contract_type"))
        contract_type = "limited" if contract_type_raw and "محدد" in contract_type_raw and "غير" not in contract_type_raw else (
            contract_type_raw if contract_type_raw in ("limited", "unlimited") else "unlimited"
        )

        employees.append(
            {
                "full_name": str(full_name).strip(),
                "employee_number": _clean_str(get(row, "employee_number")),
                "nationality": _clean_str(get(row, "nationality")),
                "job_title": _clean_str(get(row, "job_title")),
                "phone": _clean_str(get(row, "phone")),
                "salary": _parse_number(get(row, "salary")),
                "iqama_number": _clean_str(get(row, "iqama_number")),
                "iqama_expiry_date": _parse_date_cell(get(row, "iqama_expiry_date")),
                "national_id": _clean_str(get(row, "national_id")),
                "hire_date": _parse_date_cell(get(row, "hire_date")),
                "basic_salary": _parse_number(get(row, "basic_salary")),
                "housing_allowance": _parse_number(get(row, "housing_allowance")) or 0.0,
                "other_allowances": _parse_number(get(row, "other_allowances")) or 0.0,
                "contract_type": contract_type,
                "contract_end_date": _parse_date_cell(get(row, "contract_end_date")),
                "probation_end_date": _parse_date_cell(get(row, "probation_end_date")),
            }
        )

    return employees, errors


def generate_template_excel() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "الموظفون"
    sheet.append(TEMPLATE_HEADERS_AR)
    sheet.append([
        "محمد أحمد", "1001", "سعودي", "محاسب أول", "0500000000", 8000, "1234567890", "2027-01-15",
        "1012345678", "2023-03-01", 6500, 1500, 0, "غير محدد المدة", "", "2023-05-30",
    ])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
